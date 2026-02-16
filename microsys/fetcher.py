# Fundemental imports
#####################################################################
from django.http import HttpResponse, HttpResponseRedirect
from django.db.models.query import QuerySet
from django.contrib import messages
from django.db.models import FileField
from io import BytesIO
import mimetypes
import openpyxl
import zipfile

# Universal Downloader
#####################################################################
def fetch_file(request, data, file_type=None):
    """
    Universal downloader for single objects, lists, or querysets.
    
    Args:
    - request: Django request object.
    - data: A single Model instance, a List of instances, or a QuerySet.
    - file_type: Optional specific field name to download (e.g., 'pdf_file'). 
                 If None or 'all', downloads all found non-empty FileFields.
    
    Returns:
    - FileResponse (single file) or Zip Response (multiple files).
    """
    
    # 1. Normalize input to a list of instances
    records = []
    if isinstance(data, QuerySet):
        records = list(data)
    elif isinstance(data, list):
        records = data
    else:
        # Assume single model instance
        records = [data]
        
    if not records:
         messages.error(request, "لا توجد سجلات للتحميل.")
         return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    # 2. Collect Files
    files_to_download = []
    
    for record in records:
        # Smart Introspection for Metadata
        meta_info = _get_record_metadata(record)
        number = meta_info['identifier']
        date_str = meta_info['date_str']
        model_name = meta_info['model_name']
            
        # Introspect for FileFields
        model_file_fields = [
            f.name for f in record._meta.get_fields() 
            if isinstance(f, FileField)
        ]
        
        for field_name in model_file_fields:
            # Filter by requested file_type
            if file_type and file_type != 'all' and field_name != file_type:
                continue
                
            file_obj = getattr(record, field_name, None)
            if file_obj and file_obj.name:
                # Construct clean partial filename
                ext = file_obj.name.split('.')[-1]
                
                # Format: ModelName_Identifier_Date_Field.ext
                filename = f"{model_name}_{number}_{date_str}_{field_name}.{ext}"
                filename = _sanitize_filename(filename)
                
                files_to_download.append({
                    'filename': filename,
                    'file': file_obj
                })

    # 3. Decision: Error, Single File, or Zip
    if not files_to_download:
        messages.error(request, "لا توجد ملفات صالحة للتحميل في السجلات المختارة.")
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
        
    if len(files_to_download) == 1:
        # Serve Single File
        target = files_to_download[0]
        return _serve_file(target['file'], target['filename'])
    else:
        # Serve Zip
        # Name zip based on first record's model and range
        first_rec_meta = _get_record_metadata(records[0])
        last_rec_meta = _get_record_metadata(records[-1])
        
        model_name = first_rec_meta['model_name']
        first_num = first_rec_meta['identifier']
        last_num = last_rec_meta['identifier']
        
        zip_name = f"{model_name}_{first_num}-{last_num}.zip"
        zip_name = _sanitize_filename(zip_name)
        return _serve_zip(files_to_download, zip_name)


def _get_record_metadata(record):
    """
    Introspects a model instance to find the best Identifier and Date.
    Returns dict: {'identifier': str, 'date_str': str, 'model_name': str}
    """
    meta = record._meta
    fields = meta.get_fields()
    
    # 1. Identifier Strategy
    # Priority: 'number' > 'code' > 'serial' > 'reference' > pk
    identifier = None
    candidates = ['number', 'code', 'serial', 'reference', 'ref']
    
    # Check candidates by name
    for name in candidates:
        if hasattr(record, name):
            val = getattr(record, name, None)
            if val:
                identifier = str(val)
                break
    
    if not identifier:
        identifier = str(record.pk)

    # 2. Date Strategy
    # Priority: 
    #   1. Field named exactly 'date' (if Date/DateTime type)
    #   2. Any DateField (preferring those without auto_now)
    #   3. Any DateTimeField
    
    from django.db.models import DateField, DateTimeField
    
    best_date = None
    
    # Check for explicit 'date' field
    try:
        date_field = meta.get_field('date')
        if isinstance(date_field, (DateField, DateTimeField)):
            best_date = getattr(record, 'date', None)
    except Exception:
        pass
        
    if not best_date:
        # Look for DateField (Business Date)
        for f in fields:
            if isinstance(f, DateField) and not isinstance(f, DateTimeField):
                val = getattr(record, f.name, None)
                if val:
                    best_date = val
                    break
    
    if not best_date:
        # Look for DateTimeField (System Timestamp)
        for f in fields:
            if isinstance(f, DateTimeField):
                 # Prefer created_at / date_joined types
                 if 'created' in f.name or 'joined' in f.name or 'applied' in f.name:
                     val = getattr(record, f.name, None)
                     if val:
                         best_date = val
                         break
        
        # If still nothing, take any DateTimeField
        if not best_date:
            for f in fields:
                if isinstance(f, DateTimeField):
                    val = getattr(record, f.name, None)
                    if val:
                        best_date = val
                        break

    date_str = best_date.strftime('%Y-%m-%d') if best_date else 'unknown_date'
    
    return {
        'identifier': identifier,
        'date_str': date_str,
        'model_name': meta.model_name
    }


def _sanitize_filename(filename):
    """Remove unsafe characters from filename."""
    return filename.replace('/', '_').replace('\\', '_').replace(':', '-').replace(' ', '_')


def _serve_file(file_obj, filename):
    """Helper: Serve a single file with correct content type."""
    try:
        content_type, _ = mimetypes.guess_type(file_obj.name)
    except Exception:
        content_type = 'application/octet-stream'
        
    if not content_type:
        content_type = 'application/octet-stream'
        
    response = HttpResponse(content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    with file_obj.open('rb') as f:
        response.write(f.read())
        
    return response


def _serve_zip(files_list, zip_filename):
    """Helper: Create and serve a zip file from list of file objects."""
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        seen_names = set()
        for item in files_list:
            fname = item['filename']
            # Ensure uniqueness in zip
            if fname in seen_names:
                base, ext = fname.rsplit('.', 1)
                fname = f"{base}_{len(seen_names)}.{ext}"
            
            seen_names.add(fname)
            
            try:
                with item['file'].open('rb') as f:
                    zf.writestr(fname, f.read())
            except Exception:
                # If a file fails to open types (e.g. missing on disk), skip it and continue
                continue
                
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
    response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
    return response


# Excel Exporter
#####################################################################
def fetch_excel(request, queryset, exclude_fields=None, hidden_fields=None, sheet_title="Excel"):
    """
    Export a queryset to Excel with Smart Hiding.
    
    Args:
    - queryset: Data to export.
    - exclude_fields: List of field names to completely omit.
    - hidden_fields: List of field names to include but hide the column.
                     (FileFields and Auto-Timestamps are automatically hidden).
    - sheet_title: Title of the worksheet.
    """
    if not queryset:
        messages.error(request, "لا توجد بيانات للتصدير.")
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    # Normalize to iterable if list passed
    data_list = queryset
    model = None
    if isinstance(queryset, QuerySet):
        model = queryset.model
    elif isinstance(queryset, list) and queryset:
        model = queryset[0].__class__
        
    if not model:
         messages.error(request, "تعذر تحديد نموذج البيانات.")
         return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

    from openpyxl.utils import get_column_letter
    from django.db.models import FileField, DateTimeField

    # Prepare Field Lists
    exclude_fields = set(exclude_fields or [])
    user_hidden_fields = set(hidden_fields or [])
    
    final_fields = [] # List of (field_name, verbose_name, is_hidden)
    
    for field in model._meta.fields:
        if field.name in exclude_fields:
            continue
            
        is_hidden = False
        
        # 1. User specified hidden
        if field.name in user_hidden_fields:
            is_hidden = True
        
        # 2. Auto-Hide FileFields/ImageFields
        elif isinstance(field, FileField):
            is_hidden = True
            
        # 3. Auto-Hide System Timestamps
        elif isinstance(field, DateTimeField):
            if field.auto_now or field.auto_now_add:
                is_hidden = True
        
        final_fields.append({
            'name': field.name,
            'verbose': field.verbose_name.title(),
            'hidden': is_hidden
        })

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title or "Export"

    # Write Header Row
    ws.append([f['verbose'] for f in final_fields])

    # Write Data Rows
    for obj in data_list:
        row = []
        for field_info in final_fields:
            name = field_info['name']
            
            # Simple attribute access
            val = getattr(obj, name, "")
            
            # Handle FileField (use url or name)
            if hasattr(val, 'url'):
                val = val.name # or val.url if you prefer full path
            
            # Convert Model Instances / Enums / Dates to string
            if val is not None:
                row.append(str(val))
            else:
                row.append("")
        ws.append(row)

    # Apply Hidden Columns
    # openpyxl columns are 1-indexed (A=1, B=2...)
    for idx, field_info in enumerate(final_fields, start=1):
        if field_info['hidden']:
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].hidden = True

    # Filename generation
    obj_count = len(data_list)
    filename = f"{model._meta.model_name}_export_{obj_count}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
